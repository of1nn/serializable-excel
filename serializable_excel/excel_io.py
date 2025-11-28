"""
Functions for reading and writing Excel files using openpyxl.
"""

from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from serializable_excel.colors import CellStyle, CellStyleApplier


def read_excel_headers(
    worksheet: Worksheet, header_row: int = 1
) -> Dict[int, str]:
    """
    Read column headers from an Excel worksheet.

    Args:
        worksheet: openpyxl worksheet object
        header_row: Row number containing headers (1-indexed)

    Returns:
        Dictionary mapping column index (1-indexed) to header name
    """
    headers = {}
    for cell in worksheet[header_row]:
        if cell.value is not None:
            headers[cell.column] = str(cell.value).strip()
    return headers


def read_excel_rows(
    worksheet: Worksheet, start_row: int = 2, max_row: Optional[int] = None
) -> List[Dict[int, Any]]:
    """
    Read data rows from an Excel worksheet.

    Args:
        worksheet: openpyxl worksheet object
        start_row: First row to read (1-indexed, typically 2 to skip header)
        max_row: Last row to read (None for all rows)

    Returns:
        List of dictionaries mapping column index (1-indexed) to cell value
    """
    rows = []
    end_row = max_row if max_row is not None else worksheet.max_row

    for row_num in range(start_row, end_row + 1):
        row_data = {}
        row = worksheet[row_num]
        for cell in row:
            if cell.column in row_data or cell.value is None:
                continue
            row_data[cell.column] = cell.value
        # Only add non-empty rows
        if row_data:
            rows.append(row_data)

    return rows


def write_excel(
    headers: Dict[str, int],
    data_rows: List[Dict[str, Any]],
    file_path: str,
    sheet_name: str = "Sheet1",
    cell_colors: Optional[Dict[Tuple[int, int], CellStyle]] = None,
):
    """
    Write data to an Excel file.

    Args:
        headers: Dictionary mapping header names to column indices (1-indexed)
        data_rows: List of dictionaries mapping header names to values
        file_path: Path where to save the Excel file
        sheet_name: Name of the worksheet
        cell_colors: Dictionary mapping (row, col) tuples to CellStyle objects
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Write headers
    for header, col_idx in headers.items():
        ws.cell(row=1, column=col_idx, value=header)

    # Write data rows
    for row_idx, row_data in enumerate(data_rows, start=2):
        for header, col_idx in headers.items():
            value = row_data.get(header)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Apply cell color if specified
            if cell_colors is not None:
                style = cell_colors.get((row_idx, col_idx))
                if style is not None:
                    CellStyleApplier.apply(cell, style)

    # Save the workbook
    wb.save(file_path)
